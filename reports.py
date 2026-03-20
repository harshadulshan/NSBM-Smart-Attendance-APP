import customtkinter as ctk
from tkinter import messagebox
import database as db
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, timedelta
import os

# ─────────────────────────────────────────
# THEME
# ─────────────────────────────────────────
NSBM_GREEN  = "#60b94b"
NSBM_DARK   = "#050d1a"
NSBM_CARD   = "#0d1f35"
NSBM_ACCENT = "#cfff5a"
NSBM_GOLD   = "#F59E0B"
WHITE       = "#FFFFFF"
GRAY        = "#AAAAAA"
RED         = "#EF4444"
PURPLE      = "#A855F7"

class ReportsFrame(ctk.CTkFrame):
    def __init__(self, parent, user):
        super().__init__(parent, fg_color=NSBM_DARK, corner_radius=0)
        self.user = user
        self.build_ui()

    def build_ui(self):

        # ── HEADER ───────────────────────
        header = ctk.CTkFrame(
            self, fg_color=NSBM_CARD,
            corner_radius=0, height=70
        )
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkLabel(
            header,
            text="📊 Reports & Analytics",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=WHITE
        ).pack(side="left", padx=25, pady=20)

        # Export button
        ctk.CTkButton(
            header,
            text="📥 Export Excel",
            height=35,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color=NSBM_ACCENT,
            hover_color="#0099BB",
            text_color="#000000",
            corner_radius=8,
            command=self.export_excel
        ).pack(side="right", padx=25, pady=17)

        # ── FILTER BAR ───────────────────
        filter_bar = ctk.CTkFrame(
            self,
            fg_color=NSBM_CARD,
            corner_radius=0,
            height=60
        )
        filter_bar.pack(fill="x")
        filter_bar.pack_propagate(False)

        # Date range
        ctk.CTkLabel(
            filter_bar,
            text="From:",
            font=ctk.CTkFont(size=12),
            text_color=GRAY
        ).pack(side="left", padx=(20, 5), pady=15)

        self.from_date = ctk.CTkEntry(
            filter_bar,
            placeholder_text="YYYY-MM-DD",
            width=120,
            height=32,
            font=ctk.CTkFont(size=12),
            fg_color="#050d1a",
            border_color=NSBM_ACCENT,
            border_width=1
        )
        self.from_date.pack(side="left", pady=15)
        self.from_date.insert(
            0, (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        )

        ctk.CTkLabel(
            filter_bar,
            text="To:",
            font=ctk.CTkFont(size=12),
            text_color=GRAY
        ).pack(side="left", padx=(10, 5), pady=15)

        self.to_date = ctk.CTkEntry(
            filter_bar,
            placeholder_text="YYYY-MM-DD",
            width=120,
            height=32,
            font=ctk.CTkFont(size=12),
            fg_color="#050d1a",
            border_color=NSBM_ACCENT,
            border_width=1
        )
        self.to_date.pack(side="left", pady=15)
        self.to_date.insert(
            0, datetime.now().strftime('%Y-%m-%d')
        )

        # Course filter
        ctk.CTkLabel(
            filter_bar,
            text="Course:",
            font=ctk.CTkFont(size=12),
            text_color=GRAY
        ).pack(side="left", padx=(15, 5), pady=15)

        courses          = db.get_all_courses()
        self.course_map  = {'All Courses': None}
        self.course_map.update({c['name']: c['id'] for c in courses})

        self.course_var = ctk.StringVar(value="All Courses")
        ctk.CTkOptionMenu(
            filter_bar,
            values=list(self.course_map.keys()),
            variable=self.course_var,
            width=200,
            height=32,
            font=ctk.CTkFont(size=12),
            fg_color="#050d1a",
            button_color=NSBM_GREEN,
            command=self.update_batches
        ).pack(side="left", pady=15)

        # Batch filter
        ctk.CTkLabel(
            filter_bar,
            text="Batch:",
            font=ctk.CTkFont(size=12),
            text_color=GRAY
        ).pack(side="left", padx=(10, 5), pady=15)

        self.batch_var = ctk.StringVar(value="All Batches")
        self.batch_dropdown = ctk.CTkOptionMenu(
            filter_bar,
            values=["All Batches"],
            variable=self.batch_var,
            width=130,
            height=32,
            font=ctk.CTkFont(size=12),
            fg_color="#050d1a",
            button_color=NSBM_GREEN,
        )
        self.batch_dropdown.pack(side="left", pady=15, padx=(0, 10))

        # Search button
        ctk.CTkButton(
            filter_bar,
            text="🔍 Search",
            height=32,
            width=100,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color=NSBM_GREEN,
            hover_color="#1B5E20",
            corner_radius=8,
            command=self.load_reports
        ).pack(side="left", pady=15)

        # ── MAIN CONTENT ─────────────────
        content = ctk.CTkFrame(
            self, fg_color="transparent"
        )
        content.pack(fill="both", expand=True, padx=20, pady=20)

        # ── STAT CARDS ───────────────────
        self.stats_frame = ctk.CTkFrame(
            content, fg_color="transparent"
        )
        self.stats_frame.pack(fill="x", pady=(0, 15))

        # ── TABS ─────────────────────────
        self.tab_view = ctk.CTkTabview(
            content,
            fg_color=NSBM_CARD,
            segmented_button_fg_color="#050d1a",
            segmented_button_selected_color=NSBM_GREEN,
            segmented_button_selected_hover_color="#1B5E20",
            text_color=WHITE,
            corner_radius=12
        )
        self.tab_view.pack(fill="both", expand=True)

        self.tab_view.add("📋 Attendance Log")
        self.tab_view.add("👤 Student Summary")
        self.tab_view.add("⚠️ Below 80%")

        # Attendance log tab
        self.log_frame = ctk.CTkScrollableFrame(
            self.tab_view.tab("📋 Attendance Log"),
            fg_color="transparent"
        )
        self.log_frame.pack(fill="both", expand=True)

        # Student summary tab
        self.summary_frame = ctk.CTkScrollableFrame(
            self.tab_view.tab("👤 Student Summary"),
            fg_color="transparent"
        )
        self.summary_frame.pack(fill="both", expand=True)

        # Below 80% tab
        self.warning_frame = ctk.CTkScrollableFrame(
            self.tab_view.tab("⚠️ Below 80%"),
            fg_color="transparent"
        )
        self.warning_frame.pack(fill="both", expand=True)

        # Load initial data
        self.load_reports()

    def update_batches(self, course_name):
        """Update batch filter."""
        course_id = self.course_map.get(course_name)
        if course_id:
            batches    = db.get_batches_by_course(course_id)
            batch_names = ["All Batches"] + [b['name'] for b in batches]
            self.batch_dropdown.configure(values=batch_names)
            self.batch_var.set("All Batches")
        else:
            self.batch_dropdown.configure(values=["All Batches"])
            self.batch_var.set("All Batches")

    def load_reports(self):
        """Load and display reports."""
        start = self.from_date.get().strip()
        end   = self.to_date.get().strip()

        records = db.get_attendance_by_date_range(start, end)

        # Filter by course
        course_id = self.course_map.get(self.course_var.get())
        if course_id:
            records = [r for r in records
                      if r.get('course_id') == course_id]

        self.current_records = records
        self.update_stats(records)
        self.load_log_tab(records)
        self.load_summary_tab(records)
        self.load_warning_tab()

    def update_stats(self, records):
        """Update stat cards."""
        for widget in self.stats_frame.winfo_children():
            widget.destroy()

        total   = len(records)
        late    = len([r for r in records
                      if 'Late' in r.get('arrival', '')])
        on_time = total - late
        students = len(set(r['student_id'] for r in records))

        stats = [
            ("📋 Total Records",  str(total),    NSBM_ACCENT),
            ("👥 Unique Students", str(students), PURPLE),
            ("🟢 On Time",         str(on_time),  "#22C55E"),
            ("🔴 Late Arrivals",   str(late),     NSBM_GOLD),
        ]

        for i, (label, value, color) in enumerate(stats):
            card = ctk.CTkFrame(
                self.stats_frame,
                fg_color=NSBM_CARD,
                corner_radius=12
            )
            card.grid(row=0, column=i, padx=6, sticky="ew")
            self.stats_frame.grid_columnconfigure(i, weight=1)

            ctk.CTkLabel(
                card,
                text=value,
                font=ctk.CTkFont(size=28, weight="bold"),
                text_color=color
            ).pack(pady=(15, 3))

            ctk.CTkLabel(
                card,
                text=label,
                font=ctk.CTkFont(size=11),
                text_color=GRAY
            ).pack(pady=(0, 15))

    def load_log_tab(self, records):
        """Load attendance log table."""
        for widget in self.log_frame.winfo_children():
            widget.destroy()

        if not records:
            ctk.CTkLabel(
                self.log_frame,
                text="No records found for selected filters",
                font=ctk.CTkFont(size=13),
                text_color=GRAY
            ).pack(pady=30)
            return

        # Headers
        headers = ["Student ID", "Name", "Course",
                   "Batch", "Date", "Time", "Arrival", "Session"]
        widths  = [110, 160, 160, 90, 100, 80, 110, 130]

        header_row = ctk.CTkFrame(
            self.log_frame,
            fg_color="#050d1a",
            corner_radius=8
        )
        header_row.pack(fill="x", pady=(0, 5))

        for header, width in zip(headers, widths):
            ctk.CTkLabel(
                header_row,
                text=header,
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color=NSBM_ACCENT,
                width=width,
                anchor="w"
            ).pack(side="left", padx=6, pady=8)

        # Rows
        for i, record in enumerate(records):
            bg = "#050d1a" if i % 2 == 0 else "#111827"
            row = ctk.CTkFrame(
                self.log_frame,
                fg_color=bg,
                corner_radius=6
            )
            row.pack(fill="x", pady=1)

            arrival_color = NSBM_GOLD if "Late" in record.get('arrival', '') else "#22C55E"

            values = [
                (record.get('student_id',   ''), WHITE),
                (record.get('student_name', ''), WHITE),
                (record.get('course_name',  ''), GRAY),
                (record.get('batch_name',   ''), GRAY),
                (record.get('date',         ''), WHITE),
                (record.get('time',         ''), WHITE),
                (record.get('arrival',      ''), arrival_color),
                (record.get('session',      ''), NSBM_ACCENT),
            ]

            for (value, color), width in zip(values, widths):
                ctk.CTkLabel(
                    row,
                    text=str(value)[:18],
                    font=ctk.CTkFont(size=11),
                    text_color=color,
                    width=width,
                    anchor="w"
                ).pack(side="left", padx=6, pady=6)

    def load_summary_tab(self, records):
        """Load student summary."""
        for widget in self.summary_frame.winfo_children():
            widget.destroy()

        if not records:
            ctk.CTkLabel(
                self.summary_frame,
                text="No records found",
                font=ctk.CTkFont(size=13),
                text_color=GRAY
            ).pack(pady=30)
            return

        # Group by student
        student_data = {}
        for r in records:
            sid = r['student_id']
            if sid not in student_data:
                student_data[sid] = {
                    'name'    : r['student_name'],
                    'course'  : r.get('course_name', ''),
                    'batch'   : r.get('batch_name',  ''),
                    'days'    : set(),
                    'late'    : 0,
                    'on_time' : 0
                }
            student_data[sid]['days'].add(r['date'])
            if 'Late' in r.get('arrival', ''):
                student_data[sid]['late'] += 1
            else:
                student_data[sid]['on_time'] += 1

        # Total days
        all_days = len(set(r['date'] for r in records))

        # Headers
        headers = ["Student ID", "Name", "Course",
                   "Days Present", "Late", "On Time", "Attendance %"]
        widths  = [110, 160, 160, 110, 70, 80, 120]

        header_row = ctk.CTkFrame(
            self.summary_frame,
            fg_color="#050d1a",
            corner_radius=8
        )
        header_row.pack(fill="x", pady=(0, 5))

        for header, width in zip(headers, widths):
            ctk.CTkLabel(
                header_row,
                text=header,
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color=NSBM_ACCENT,
                width=width,
                anchor="w"
            ).pack(side="left", padx=6, pady=8)

        for i, (sid, data) in enumerate(student_data.items()):
            days_present = len(data['days'])
            percentage   = round((days_present/all_days)*100, 1) if all_days > 0 else 0
            pct_color    = RED if percentage < 80 else "#22C55E"

            bg = "#050d1a" if i % 2 == 0 else "#111827"
            row = ctk.CTkFrame(
                self.summary_frame,
                fg_color=bg,
                corner_radius=6
            )
            row.pack(fill="x", pady=1)

            values = [
                (sid,                 WHITE),
                (data['name'],        WHITE),
                (data['course'],      GRAY),
                (str(days_present),   NSBM_ACCENT),
                (str(data['late']),   NSBM_GOLD),
                (str(data['on_time']),  "#22C55E"),
                (f"{percentage}%",    pct_color),
            ]

            for (value, color), width in zip(values, widths):
                ctk.CTkLabel(
                    row,
                    text=str(value),
                    font=ctk.CTkFont(size=11),
                    text_color=color,
                    width=width,
                    anchor="w"
                ).pack(side="left", padx=6, pady=6)

    def load_warning_tab(self):
        """Load students below 80% attendance."""
        for widget in self.warning_frame.winfo_children():
            widget.destroy()

        ctk.CTkLabel(
            self.warning_frame,
            text="⚠️ Students Below 80% Attendance",
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=RED
        ).pack(anchor="w", pady=(10, 15))

        students = db.get_all_students()
        warnings = []

        for student in students:
            pct = db.get_student_attendance_percentage(
                student['student_id']
            )
            if pct < 80:
                warnings.append((student, pct))

        if not warnings:
            ctk.CTkLabel(
                self.warning_frame,
                text="✅ All students have attendance above 80%!",
                font=ctk.CTkFont(size=13),
                text_color="#22C55E"
            ).pack(pady=30)
            return

        for student, pct in sorted(warnings, key=lambda x: x[1]):
            card = ctk.CTkFrame(
                self.warning_frame,
                fg_color="#2D0000",
                corner_radius=10
            )
            card.pack(fill="x", pady=4)

            ctk.CTkLabel(
                card,
                text=f"⚠️  {student['student_id']} — {student['full_name']}",
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color=RED
            ).pack(side="left", padx=15, pady=12)

            ctk.CTkLabel(
                card,
                text=student.get('course_name', ''),
                font=ctk.CTkFont(size=11),
                text_color=GRAY
            ).pack(side="left", padx=5)

            ctk.CTkLabel(
                card,
                text=f"{pct}%",
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color=RED
            ).pack(side="right", padx=15)

    def export_excel(self):
        """Export report to Excel."""
        if not hasattr(self, 'current_records') or \
           not self.current_records:
            messagebox.showwarning(
                "Warning", "No records to export!"
            )
            return

        os.makedirs('reports', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        path      = f'reports/attendance_report_{timestamp}.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Title
        ws.merge_cells('A1:H1')
        title           = ws['A1']
        title.value     = f"NSBM Smart Attendance Report — {datetime.now().strftime('%B %d, %Y')}"
        title.font      = Font(bold=True, size=14, color='00D9FF')
        title.fill      = PatternFill(start_color='0D1117', end_color='0D1117', fill_type='solid')
        title.alignment = Alignment(horizontal='center')

        # Headers
        headers = ["Student ID", "Name", "Course", "Batch",
                  "Date", "Time", "Arrival", "Session"]
        for col, header in enumerate(headers, 1):
            cell            = ws.cell(row=2, column=col, value=header)
            cell.font       = Font(color='00D9FF', bold=True)
            cell.fill       = PatternFill(start_color='161B22', end_color='161B22', fill_type='solid')
            cell.alignment  = Alignment(horizontal='center')

        # Data
        for row, record in enumerate(self.current_records, 3):
            ws.cell(row=row, column=1, value=record.get('student_id',   ''))
            ws.cell(row=row, column=2, value=record.get('student_name', ''))
            ws.cell(row=row, column=3, value=record.get('course_name',  ''))
            ws.cell(row=row, column=4, value=record.get('batch_name',   ''))
            ws.cell(row=row, column=5, value=record.get('date',         ''))
            ws.cell(row=row, column=6, value=record.get('time',         ''))
            ws.cell(row=row, column=7, value=record.get('arrival',      ''))
            ws.cell(row=row, column=8, value=record.get('session',      ''))

        # Column widths
        widths = [15, 25, 30, 15, 12, 10, 15, 20]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = width

        wb.save(path)
        messagebox.showinfo(
            "Success",
            f"✅ Report exported to:\n{path}"
        )